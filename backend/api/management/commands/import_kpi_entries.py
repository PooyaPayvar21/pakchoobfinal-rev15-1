from decimal import Decimal
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from api.models import KPIEntry

class Command(BaseCommand):
    help = "Import KPI entries from an Excel file into KPIEntry table"

    def add_arguments(self, parser):
        parser.add_argument("--file", type=str, required=True)
        parser.add_argument("--sheet", type=str, default="Sheet1")
        parser.add_argument("--season", type=str, default="")

    def handle(self, *args, **options):
        file_path = options.get("file")
        sheet_name = options.get("sheet")
        season = options.get("season")

        try:
            import openpyxl
        except Exception as e:
            self.stdout.write(self.style.ERROR("Missing dependency: openpyxl"))
            self.stdout.write("Install with: pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(filename=file_path, data_only=True)
            ws = wb[sheet_name]
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error opening workbook: {str(e)}"))
            return

        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h or "").strip() for h in header_cells]

        def norm(s):
            return str(s).strip().lower().replace("_", " ")

        header_map = {}
        for idx, h in enumerate(headers):
            key = norm(h)
            if key in ("company_name", "company name", "نام شرکت"):
                header_map[idx] = "company_name"
            elif key in ("season", "فصل"):
                header_map[idx] = "season"
            elif key in ("personal_code", "personal code", "کد پرسنلی"):
                header_map[idx] = "personal_code"
            elif key in ("full_name", "full name", "نام و نام خانوادگی"):
                header_map[idx] = "full_name"
            elif key in ("role", "عنوان شغلی"):
                header_map[idx] = "role"
            elif key in ("direct_management", "direct management", "مدیر مستقیم"):
                header_map[idx] = "direct_management"
            elif key in ("departman", "department", "دپارتمان"):
                header_map[idx] = "departman"
            elif key in ("category", "دسته بندی"):
                header_map[idx] = "category"
            elif key in ("obj weight", "object weight"):
                header_map[idx] = "obj_weight"
            elif key in ("kpi en", "kpien"):
                header_map[idx] = "kpi_en"
            elif key in ("kpi fa", "kpifa"):
                header_map[idx] = "kpi_fa"
            elif key in ("kpi info", "kpi information"):
                header_map[idx] = "kpi_info"
            elif key in ("target"):
                header_map[idx] = "target"
            elif key in ("kpi weight", "kpiweight"):
                header_map[idx] = "kpi_weight"
            elif key in ("kpi achievement", "kpiachievement"):
                header_map[idx] = "kpi_achievement"
            elif key in ("percentage achievement", "percentage"):
                header_map[idx] = "score_achievement_alt"
            elif key in ("score achievement", "score"):
                header_map[idx] = "score_achievement"
            elif key in ("type"):
                header_map[idx] = "entry_type"
            elif key in ("sum", "total"):
                header_map[idx] = "sum_value"

        created = 0
        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {}
            for idx, value in enumerate(row):
                field = header_map.get(idx)
                if not field:
                    continue
                if field in (
                    "obj_weight",
                    "target",
                    "kpi_weight",
                    "kpi_achievement",
                    "score_achievement",
                    "score_achievement_alt",
                    "sum_value",
                ):
                    if value is None or value == "":
                        data[field] = None
                    else:
                        try:
                            data[field] = Decimal(str(value))
                        except Exception:
                            data[field] = None
                else:
                    data[field] = str(value).strip() if value is not None else ""

            if season and not data.get("season"):
                data["season"] = season

            personal_code = data.get("personal_code", "")
            kpi_en = data.get("kpi_en", "")

            if not personal_code and not kpi_en:
                skipped += 1
                continue

            try:
                existing = None
                if personal_code and kpi_en:
                    existing = KPIEntry.objects.filter(
                        personal_code=personal_code, kpi_en=kpi_en
                    ).first()

                if existing:
                    for k, v in data.items():
                        setattr(existing, k, v)
                    existing.save()
                    updated += 1
                else:
                    KPIEntry.objects.create(**data)
                    created += 1
            except Exception as e:
                skipped += 1

        self.stdout.write(self.style.SUCCESS(f"Created: {created}"))
        self.stdout.write(self.style.SUCCESS(f"Updated: {updated}"))
        self.stdout.write(self.style.WARNING(f"Skipped: {skipped}"))
