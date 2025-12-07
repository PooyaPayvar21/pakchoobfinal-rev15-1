from django.core.management.base import BaseCommand, CommandError
from api.models import KPIPersonel


class Command(BaseCommand):
    help = "Import KPIPersonel rows from an Excel sheet"

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            default=r"d:\Pooya\Project\pakchoobfinal-rev15-1\Copy of روال تاییدات kpi.xlsx",
        )
        parser.add_argument("--sheet", type=str, default="Group")
        parser.add_argument("--truncate", action="store_true")

    def handle(self, *args, **options):
        try:
            import openpyxl
        except Exception as e:
            raise CommandError("openpyxl is required: pip install openpyxl")

        path = options["path"]
        sheet_name = options["sheet"]

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            raise CommandError(f"Unable to open Excel file: {e}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        headers_row = None
        for row in ws.iter_rows(values_only=True):
            if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                headers_row = [str(cell).strip() if cell is not None else "" for cell in row]
                break
        if not headers_row:
            raise CommandError("No header row detected")

        normalized = [h.replace("\u200c", "").strip() for h in headers_row]

        def idx(name):
            try:
                return normalized.index(name)
            except ValueError:
                return None

        idx_full_name = idx("نام و نام خانوادگی")
        idx_personal_code = idx("کد پرسنلی")
        idx_departman = idx("معاونت")
        idx_job_title = idx("عنوان شغلی")

        required = {
            "نام و نام خانوادگی": idx_full_name,
            "کد پرسنلی": idx_personal_code,
            "معاونت": idx_departman,
            "عنوان شغلی": idx_job_title,
        }
        missing = [k for k, v in required.items() if v is None]
        if missing:
            raise CommandError(f"Missing required columns: {', '.join(missing)}")

        if options["truncate"]:
            KPIPersonel.objects.all().delete()

        created = 0
        updated = 0
        rows = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            rows += 1
            full_name = row[idx_full_name] if row[idx_full_name] is not None else ""
            personal_code = row[idx_personal_code] if row[idx_personal_code] is not None else ""
            departman = row[idx_departman] if row[idx_departman] is not None else ""
            job_title = row[idx_job_title] if row[idx_job_title] is not None else ""

            full_name = str(full_name).strip()
            personal_code = str(personal_code).strip()
            departman = str(departman).strip()
            job_title = str(job_title).strip()

            if not personal_code:
                continue

            obj, is_created = KPIPersonel.objects.update_or_create(
                personal_code=personal_code,
                defaults={
                    "full_name": full_name,
                    "departman": departman,
                    "job_title": job_title,
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Processed {rows} data rows. Created: {created}, Updated: {updated}"
            )
        )
