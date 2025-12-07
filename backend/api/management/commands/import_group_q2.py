from django.core.management.base import BaseCommand, CommandError
from decimal import Decimal, InvalidOperation
from api.models import GroupQ2Entry
import re


class Command(BaseCommand):
    help = "Import GROUPQ2 rows from an Excel sheet into GroupQ2Entry"

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            default=r"d:\Pooya\Project\pakchoobfinal-rev15-1\q2-1404.xlsx",
        )
        parser.add_argument("--sheet", type=str, default="GROUPQ2")
        parser.add_argument("--truncate", action="store_true")

    def handle(self, *args, **options):
        try:
            import openpyxl
        except Exception:
            raise CommandError("openpyxl is required: pip install openpyxl")

        path = options["path"]
        sheet_name = options["sheet"]

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            raise CommandError(f"Unable to open Excel file: {e}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        headers = None
        for row in ws.iter_rows(values_only=True):
            if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                break
        if not headers:
            raise CommandError("No header row detected")

        norm_headers = [re.sub(r"\s+", " ", h.replace("\u200c", "").strip().lower()) for h in headers]

        def find_index(names):
            for name in names:
                try:
                    return norm_headers.index(name)
                except ValueError:
                    continue
            return None

        idx_company = find_index(["company_name", "شرکت", "شرکت/واحد", "facility", "گروه صنعتی"])
        idx_season = find_index(["season", "فصل"])
        idx_personal = find_index(["personal_code", "کد پرسنلی", "کدپرسنلی"])
        idx_fullname = find_index(["full_name", "نام و نام خانوادگی", "نام نام خانوادگی", "نام"])
        idx_job_title = find_index(["job_title", "عنوان شغلی", "سمت"])
        idx_manager_code = find_index([
            "direct_manager_code",
            "manager_code",
            "کد پرسنلی مدیر مستقیم",
            "کد مدیر",
            "کد مدیر مستقیم",
        ])
        idx_manager_name = find_index([
            "manager_name",
            "نام مدیر مستقیم",
            "نام مدیر",
            "مدیر مستقیم",
        ])
        idx_departman = find_index(["departman", "معاونت", "دپارتمان", "واحد سازمانی"])
        idx_category_fa = find_index([
            "category_fa",
            "گروه کاری",
            "دسته",
            "گروه",
            "دسته بندی",
            "دستهبندی",
        ])
        idx_category_en = find_index(["category_en", "category", "group"])
        idx_obj_weight = find_index(["obj_weight", "وزن هدف", "وزن"])
        idx_kpi_en = find_index([
            "kpi_en",
            "kpi en",
            "kpi english",
            "kpi (en)",
            "kpi-en",
            "english kpi",
            "kpi title en",
        ])
        idx_kpi_fa = find_index([
            "kpi_fa",
            "kpi fa",
            "kpi",
            "شاخص",
            "شرح شاخص",
            "عنوان شاخص",
            "kpi (fa)",
            "kpi-fa",
        ])
        idx_kpi_info = find_index([
            "kpi_info",
            "kpi info",
            "info",
            "توضیحات",
            "شرح",
            "notes",
            "kpi details",
            "شرح شاخص",
        ])
        idx_target = find_index(["target", "هدف", "target value"])
        idx_kpi_weight = find_index(["kpi_weight", "وزن شاخص"])
        idx_kpi_achievement = find_index(["kpi_achievement", "تحقق", "achievement"])
        idx_percentage_ach = find_index(["percentage_achievement", "درصد تحقق", "percentage"])
        idx_score_ach = find_index(["score_achievement", "امتیاز تحقق", "score"])
        idx_entry_type = find_index(["entry_type", "نوع", "type", "وضعیت"])
        idx_sum_percent = find_index(["sum_percent", "sum", "جمع درصد", "جمع"])
        idx_notes = find_index(["notes", "یادداشت", "توضیحات"])

        if options["truncate"]:
            GroupQ2Entry.objects.all().delete()

        def to_decimal(v):
            if v is None:
                return None
            s = str(v).strip().replace("%", "").replace(",", "")
            if s == "":
                return None
            try:
                val = Decimal(s)
            except InvalidOperation:
                return None
            if val.copy_abs() >= Decimal("1e10"):
                return None
            return val

        def to_percent_int(v):
            """Parse a percentage-like value and return an integer 0-100 as Decimal.

            Examples:
            - "100%" -> 100
            - "90" -> 90
            - 90.4 -> 90
            - "0.9" -> 1 (no fractional scaling)
            """
            val = to_decimal(v)
            if val is None:
                return None
            try:
                n = int(round(float(val)))
            except Exception:
                n = 0
            if n < 0:
                n = 0
            if n > 100:
                n = 100
            return Decimal(n)

        created = 0
        updated = 0
        rows = 0
        row_index_counter = 1

        if idx_kpi_en is None and idx_kpi_fa is None and idx_kpi_info is None:
            self.stdout.write(self.style.WARNING(f"KPI columns not found in headers: {headers}"))

        for row in ws.iter_rows(min_row=2, values_only=True):
            rows += 1
            company = row[idx_company] if idx_company is not None else ""
            season = row[idx_season] if idx_season is not None else ""
            personal_code = row[idx_personal] if idx_personal is not None else ""
            full_name = row[idx_fullname] if idx_fullname is not None else ""
            job_title = row[idx_job_title] if idx_job_title is not None else ""
            manager_code = row[idx_manager_code] if idx_manager_code is not None else ""
            manager_name = row[idx_manager_name] if idx_manager_name is not None else ""
            departman = row[idx_departman] if idx_departman is not None else ""
            category_fa = row[idx_category_fa] if idx_category_fa is not None else ""
            category_en = row[idx_category_en] if idx_category_en is not None else ""
            obj_weight = to_decimal(row[idx_obj_weight] if idx_obj_weight is not None else None)
            kpi_en = row[idx_kpi_en] if idx_kpi_en is not None else ""
            kpi_fa = row[idx_kpi_fa] if idx_kpi_fa is not None else ""
            kpi_info = row[idx_kpi_info] if idx_kpi_info is not None else ""
            target = to_decimal(row[idx_target] if idx_target is not None else None)
            kpi_weight = to_decimal(row[idx_kpi_weight] if idx_kpi_weight is not None else None)
            kpi_achievement = to_decimal(row[idx_kpi_achievement] if idx_kpi_achievement is not None else None)
            percentage_achievement = to_percent_int(row[idx_percentage_ach] if idx_percentage_ach is not None else None)
            score_achievement = to_decimal(row[idx_score_ach] if idx_score_ach is not None else None)
            entry_type = row[idx_entry_type] if idx_entry_type is not None else ""
            sum_percent = to_percent_int(row[idx_sum_percent] if idx_sum_percent is not None else None)
            notes = row[idx_notes] if idx_notes is not None else ""

            def to_str(v):
                return "" if v is None else str(v).strip()

            def limit(s, n):
                return s[:n] if len(s) > n else s

            company = limit(to_str(company), 200)
            season = limit(to_str(season), 50)
            personal_code = limit(to_str(personal_code), 50)
            full_name = limit(to_str(full_name), 200)
            job_title = limit(to_str(job_title), 100)
            manager_code = limit(to_str(manager_code), 50)
            manager_name = limit(to_str(manager_name), 200)
            departman = limit(to_str(departman), 200)
            category_fa = limit(to_str(category_fa), 200)
            category_en = limit(to_str(category_en), 200)
            kpi_en = limit(to_str(kpi_en), 250)
            kpi_fa = limit(to_str(kpi_fa), 250)
            kpi_info = to_str(kpi_info)
            entry_type = limit(to_str(entry_type), 100)
            notes = to_str(notes)

            if personal_code == "" and full_name == "":
                row_index_counter += 1
                continue

            lookup = {}
            if personal_code and kpi_fa and season:
                lookup = {"personal_code": personal_code, "kpi_fa": kpi_fa, "season": season}
            else:
                lookup = {"row_index": row_index_counter}

            obj, is_created = GroupQ2Entry.objects.update_or_create(
                **lookup,
                defaults={
                    "row_index": row_index_counter,
                    "company_name": company,
                    "season": season,
                    "personal_code": personal_code,
                    "full_name": full_name,
                    "job_title": job_title,
                    "direct_manager_code": manager_code,
                    "manager_name": manager_name,
                    "departman": departman,
                    "category_fa": category_fa,
                    "category_en": category_en,
                    "obj_weight": obj_weight,
                    "kpi_en": kpi_en,
                    "kpi_fa": kpi_fa,
                    "kpi_info": kpi_info,
                    "target": target,
                    "kpi_weight": kpi_weight,
                    "kpi_achievement": kpi_achievement,
                    "percentage_achievement": percentage_achievement,
                    "score_achievement": score_achievement,
                    "entry_type": entry_type,
                    "sum_percent": sum_percent,
                    "notes": notes,
                },
            )

            if is_created:
                created += 1
            else:
                updated += 1

            row_index_counter += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Processed {rows} data rows. Created: {created}, Updated: {updated}"
            )
        )
